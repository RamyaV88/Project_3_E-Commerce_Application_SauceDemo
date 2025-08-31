import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value


def write_data(file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file)


def get_data_from_excel(file, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for row_num in range(2, total_rows + 1):
        row_list = []
        for col_num in range(1, total_cols + 1):
            row_list.append(sheet.cell(row=row_num, column=col_num).value)
        final_list.append(row_list)
    return final_list
